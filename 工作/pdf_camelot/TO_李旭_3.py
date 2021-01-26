# coding: utf-8
# need install modules：camelot-py、opencv-python、pdfplumber、pandas、openpyxl
# reference doc：https://github.com/jsvine/pdfplumber

import pdfplumber
import os.path as ps
import re
from openpyxl import Workbook

# glob 英文：水滴，一滴，一团。 unix，php，pyhton都有对glob的实现。
# * 匹配任意 0 或多个任意字符
# ? 匹配任意一个字符
# [] 若字符在中括号中，例如[0-9]匹配数字
import glob


# pd.set_option('display.max_columns', 1000)
# pd.set_option('display.width', 1000)
# pd.set_option('display.max_colwidth', 1000)


def parse_pdf(file_path, ws, start_row):
    print('---------------------提取表格 Start--------------------------')
    pdf = pdfplumber.open(file_path)
    final_arr = []
    del_flag = False
    # 更改前加入了表格，那么下一个更改后也加入表格，防止出现多个更改前更改后，而有些是跟着图片非表格，不需要加入表格的，所以按更改前更改后成对匹配表格
    change_flag = 0
    # 读取表格时，由于优先读取包含表头的，若表头上面还有无表头表格，当有表头的加入列表后，再次读到无表头的需要设置到当前Index-2的位置，为不是-1
    for p in range(0, len(pdf.pages)):
        table_lst = []
        text_lst = []
        page = pdf.pages[p]
        page_text = page.extract_text()

        # 从第一页提取ABC列
        if p == 0:
            # 文件名
            str_a = ps.basename(file_path).replace('.pdf', '')
            print('【A' + str(start_row) + '】: ', str_a)

            # 更改控制编号
            str_b = re.compile(r'更改控制编号\s{2,}([A-Za-z0-9_]+\s*[A-Za-z0-9_]+)\s*').findall(page_text)[0].replace(' ', '')
            print('【B' + str(start_row) + '】: ', str_b)

            # 更改执行架次
            str_c = re.compile(r'更改执行架次\s{2,}(.+)\.').findall(page_text)[0].replace(' ', '')
            print('【C' + str(start_row) + '】: ', str_c)

            # 写入A、B、C列
            ws.cell(row=start_row, column=1, value=str_a)
            ws.cell(row=start_row, column=2, value=str_b)
            ws.cell(row=start_row, column=3, value=str_c)

        # 从包含“更改内容"的页面开始处理
        if not del_flag:
            if '更改内容' in page_text:
                del_flag = True
                page_text = page_text[page_text.find('更改内容', 0):]
            else:
                continue

        page_tables = page.extract_tables(
            table_settings={'vertical_strategy': 'lines_strict', 'horizontal_strategy': 'lines_strict',
                            'min_words_vertical': 1, 'min_words_horizontal': 1})

        # 重置每一页实际提取表格个数
        pt_flag = 0

        # 当页无表头表格所在index
        no_title_idx = 0
        for pt in range(0, len(page_tables)):
            page_table = page_tables[pt]
            if page_table[0][0] != '' and page_table[0][0] != '线束号' and page_table[0][0].strip().replace('\n', '') != '设备号(从)':
                no_title_idx = pt
                break

        for pt in range(0, len(page_tables)):
            page_table = page_tables[pt]
            # 列数= [10, 11, 12, 13, 14] 中一种的才处理，也就是过滤掉不相关的小表格
            if page_table and len(page_table[0]) in [9, 10, 11, 12, 13, 14]:
                pt_flag += 1
                # print('【table】↓\n', page_table)
                # 1.只要线束号或设备号(从)开头的，其它的过滤掉
                # 2.有些PDF中，表格换页了，但是下一页接着的表格是带表头的，需要判断是不是要拼接到上一个表中
                #   目前判断逻辑：判断当页text是不是以线束号、或者设备号(从)开头，是就拼接到上一个，不是就新增
                temp_arr = page_text.split('\n')
                # 开始内容(读取的前三行是新型涡扇支线飞机项目、工程指令、系列更改之类的，第四行开始内容)
                start_row_content = temp_arr[3]
                # 如果开始内容不包含符号，则是表表格开始(拼接到上一个)，如果包含符号则是文字开始(新增表格)
                if ('更改内容' in temp_arr[0] or pt > 1
                    or re.match('.*[，。：；！、”]', start_row_content)
                    # 如果当页有多个表格，并且有那种无表头的在非第2个(index=1)表格中(特殊情况会读取不按PDF中顺序读两个表；第一个表格是工程指令那个固定模板，第二个才开始内容中的表格)
                    or (len(page_tables) > 1 and no_title_idx >= 1)
                ) and (page_table[0][0] == '线束号' or page_table[0][0].strip().replace('\n', '') == '设备号(从)'):
                    table_lst.append(page_table)
                    # print('ok, 带表头的独立表格，加入表格成功!')
                else:
                    # 先把多余的表头行去掉(这里的表格都是要追加到上一个表格的，有的带表头有的不带，带就要去掉)
                    if page_table[0][0] == '线束号' or page_table[0][0].strip().replace('\n', '') == '设备号(从)':
                        page_table.pop(0)

                    # 不含表头的拼接至上一个带表头的数组中，可能是分页造成的数据断裂
                    for idx in range(len(final_arr) - 1, -1, -1):
                        # 倒序循环找到上一个表格
                        if isinstance(final_arr[idx], list):
                            if idx == len(final_arr) - 1:
                                final_arr += page_table
                            else:
                                final_arr = final_arr[0:idx] + page_table + final_arr[idx + 1:len(final_arr)]
                            break

                    # last_index = len(table_lst) - table_insert_flag
                    # # 断裂数据也要过滤掉不相关的数据：不是以线束号、设备号(从)开头的不要
                    # if last_index >= 0 and (table_lst[last_index][0][0] == '线束号' or
                    #                         page_table[0][0].strip().replace('\n', '') == '设备号(从)'):
                    #     table_lst[last_index] = table_lst[last_index] + page_table
                    # print('ok, 无表头，追加到上一个表格中!')
            else:
                arr_str = page_text.split('\n')
                arr_new = [s for s in arr_str if not s.startswith('W')
                           and not s.startswith('新型涡扇支线飞机项目')
                           and not s.startswith('工程指令')
                           and not s.startswith('EO_253A')
                           and not s.startswith('线束号 ')
                           and not s.startswith('EO')
                           and s not in ['ATA', '编号', 'ATA编号']
                           and (len(s.split(' ')) <= 3 or re.match('.*[，。：；！、”]', s))
                           and not re.match(r'.+系列更改.+', s)
                           and not re.match(r'第.+页.+共.+页', s)
                           and not re.match(r'第.+页.+共.+页', s)
                           and not (len(s) == 3 and s.startswith('格式'))
                           and s not in ['导 孔', '端接 端接', '料 )', '导料', '(从) (到)']
                           ]
                if arr_new:
                    # 若上一行文字结尾不是以：，。结尾，则需要拼接在上一行中，是因为pdf中换行了，分割开了
                    arr_new_temp = []
                    exist_len = 0
                    for a in range(0, len(arr_new)):
                        t_str = arr_new[a]
                        if a == 0 or re.match(r'^\d+\).*$', t_str) or re.match(r'^\d+）.*$', t_str) \
                                or arr_new_temp[exist_len - 1].endswith('，') or arr_new_temp[
                            exist_len - 1].endswith('。') or arr_new_temp[exist_len - 1].endswith('：') \
                                or arr_new_temp[exist_len - 1].endswith('；') or arr_new_temp[
                            exist_len - 1].endswith('！') or arr_new_temp[exist_len - 1].endswith('”'):
                            arr_new_temp.append(t_str)
                            exist_len += 1
                        else:
                            arr_new_temp[exist_len - 1] = arr_new_temp[exist_len - 1] + t_str

                    # 再次筛选数据
                    arr_new_temp = [s for s in arr_new_temp if re.match('.*[，。：；！、”]', s)]
                    text_lst += arr_new_temp

        print('【text】 len=', len(text_lst), '↓\n', text_lst)
        print('【table】 len=', len(table_lst), '↓\n', table_lst)
        result_arr = []

        for t in range(0, len(text_lst)):
            text = text_lst[t]
            if text.strip().startswith('更改内容'):
                result_arr.append(text)
                continue
            if text.endswith('：') or text.endswith('如下表所示。'):
                # text以冒号结尾，则拼接表格数据, 防止更改前连续两行冒号的情况，需要判断下一行是不是更改前：结束
                # 并且下一行不是以 2) xxx 这样的开头
                if len(table_lst) > 0 and ((t != len(text_lst) - 1 and not text_lst[t + 1].endswith('更改前：')
                                            and '图纸' not in text)
                                           or t == len(text_lst) - 1):
                    if text != '更改后：' or (text == '更改后：' and change_flag == 1):
                        table_data = table_lst.pop(0)
                        result_arr.append(text)
                        for data in table_data:
                            result_arr.append(data)

                        if text == '更改前：':
                            # 更改前加入了表格，那么下个更改后无需判断直接加入表格
                            change_flag = 1
                        elif text == '更改后：':
                            # 更改后加入了表格，那么重置状态，下一个更改前要重新判断
                            change_flag = 0
                    else:
                        result_arr.append(text)
                else:
                    result_arr.append(text)
            else:
                # 不需要拼接表格数据的，
                result_arr.append(text)

        final_arr += result_arr
    print('【提取内容】 len=', len(final_arr), '↓\n', final_arr)

    # 写入更改内容（包括文字和表格）
    for r in range(start_row, len(final_arr) + start_row):
        row_val = final_arr[r - start_row]
        if isinstance(row_val, list):
            # 是表格数据，则按列写入
            w_col = 4
            for d in row_val:
                ws.cell(row=r, column=w_col, value=d)
                w_col += 1
        else:
            # 文本直接单元格写入
            ws.cell(row=r, column=4, value=row_val)
    print('---------------------提取表格 End--------------------------')
    return [ws, start_row + len(final_arr)]


if __name__ == '__main__':
    # folder_path =  r'../*.py'
    folder_path =  r'.\PDF\*.pdf'
    save_excel_path = r'./result.xlsx'

    pdf_arr = glob.glob(folder_path)
    print(folder_path, pdf_arr)

    # 新建 Excel
    wb_new = Workbook()
    ws_new = wb_new.active

    # 追加标题
    ws_new.append(['EO编号', 'ECP/DCR', '更改架次'])

    next_row = 2

    # 循环读取pdf
    for pdf_path in pdf_arr:
        print('开始处理pdf: ', pdf_path)
        ret_arr = parse_pdf(file_path=pdf_path, ws=ws_new, start_row=next_row)
        ws_new = ret_arr[0]
        next_row = ret_arr[1]

        # 保存
        wb_new.save(save_excel_path)
        wb_new.close()
